import requests
import json
import jsonpath
import openpyxl
from openpyxl.utils import column_index_from_string


def test_add_Multiple_data():
    BaseUrl = "https://thetestingworldapi.com/api/studentsDetails"
    f = open('C:\\Soumyaranjan_Personal\\Study\\API_Testing\\TestCases\\AddStudent.json', 'r')
    TestData = openpyxl.load_workbook("C:\\Soumyaranjan_Personal\\Study\\API_Testing\\TestCases\\TestData.xlsx")
    sheetName = TestData['Sheet1']
    rows = sheetName.max_row
    jsonRequest = json.loads(f.read())

    for i in range(2, rows+1):
        cell_first_name = sheetName.cell(row=i, column=1)
        cell_middle_name =  sheetName.cell(row=i, column=2)
        cell_last_name = sheetName.cell(row=i, column=3)
        cell_dob = sheetName.cell(row=i, column=4)

        jsonRequest['first_name']= cell_first_name.value
        jsonRequest['middle_name']= cell_middle_name.value
        jsonRequest['last_name']= cell_last_name.value
        jsonRequest['date_of_birth']= cell_dob.value


    response = requests.post(BaseUrl, jsonRequest)
    print(response.text)
    print(response.status_code)
    assert response.status_code == 201