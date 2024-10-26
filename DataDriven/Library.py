import json
import jsonpath
import requests
import openpyxl

class Commom:

    def __init__(self, FileNamePath, SheetName):
        global TestData
        global sheetName
        TestData = openpyxl.load_workbook(FileNamePath)
        sheetName = TestData[SheetName]

    def fetch_row_count(self):
        rows = sheetName.max_row
        return rows

    def fetch_column_count(self):
        column = sheetName.max_column
        return column

    def fetch_key_names(self):
        c = sheetName.max_column
        li=[]
        for i in range(1,c+1):
            cell = sheetName.cell(row=1, column=i)
            li.insert(i-1,cell.value)
        return li

    def update_request_with_data(self, rowNumber,jsonRequest,keyList):
        c = sheetName.max_column
        for i in range(1, c+1):
            cell = sheetName.cell(row=rowNumber, column=i)
            jsonRequest[keyList[i-1]] = cell.value

        return jsonRequest